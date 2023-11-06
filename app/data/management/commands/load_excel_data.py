from django.core.management.base import BaseCommand
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment
from data.models import Samochod


def generate_rating():
    return random.randint(0, 4)


class Command(BaseCommand):
    help = 'Load sample data into the Excel Sheet'
    default_rows_amount = 10
    file_name = "stan_techniczny_samochodow.xlsx"

    def add_arguments(self, parser):
        pass

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS(f'Loading data to Excel Sheet ...'))

        wb = Workbook()
        ws = wb.active
        ws.title = "Stan Techniczny Samochodow"

        columns = [
            "Numer rejestracyjny",
            "Przebieg",
            "Poziom oleju",
            "Poziom płynu chłodniczego",
            "Poziom płynu hamulcowego",
            "Poziom płynu do spryskiwaczy",
            "Stan wizualny",
            "Stan techniczny"
        ]

        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.alignment = Alignment(horizontal="center")

        cars = Samochod.objects.all()
        for row_num in range(2, Samochod.objects.count() + 2):
            car = cars[row_num-2]
            ws.cell(row=row_num, column=1, value=car.numer_rejestracyjny)
            ws.cell(row=row_num, column=2, value=random.randint(1000, 50000))
            ws.cell(row=row_num, column=3, value=generate_rating())
            ws.cell(row=row_num, column=4, value=generate_rating())
            ws.cell(row=row_num, column=5, value=generate_rating())
            ws.cell(row=row_num, column=6, value=generate_rating())
            ws.cell(row=row_num, column=7, value=generate_rating())
            ws.cell(row=row_num, column=8, value=generate_rating())

        wb.save(self.file_name)
        self.stdout.write(self.style.SUCCESS(f'Excel data saved to {self.file_name}'))

