import pickle
from typing import List
import os
from app.settings import BACKUP_DIR
from .models import Pracownik, Samochod, Klient, Oddzial, Wypozyczenie
from django.db.models import Model
import logging
import random
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import date
from app import settings
import shutil
from faker import Factory as FakerFactory

faker = FakerFactory.create('pl_PL')


def generate_rating():
    return random.randint(0, 4)


logger = logging.getLogger(__name__)


def get_default_models():
    return [Oddzial, Pracownik, Samochod, Klient, Wypozyczenie]


def remove(model_class: Model, amount: int):
    model_class.objects.all().delete()
    logger.info(f'{amount} {model_class.__name__} instances removed successfully')


def bulk_create(model_class: Model, instances: List[Model]) -> None:
    model_class.objects.bulk_create(instances)
    logger.info(f'{model_class.__name__} instances created successfully')


def clear_models(*models):
    if not models:
        models = get_default_models()
    for model, amount in [(model, model.objects.count()) for model in models]:
        remove(model, amount)


def create_database_backup(*models):
    if not models:
        models = get_default_models()
    today = date.today()
    backup_dir = BACKUP_DIR / 'database' / str(today)
    os.makedirs(backup_dir, exist_ok=True)
    for model in models:
        filename = f'{model.__name__.lower()}_backup.pkl'
        backup_path = backup_dir / filename
        backup = model.objects.all()
        with open(backup_path, 'wb') as file:
            pickle.dump(backup, file)


def load_database_backup(*models, date=None):
    if not models:
        models = get_default_models()

    if date is None:
        backup_dirs = [d for d in os.listdir(BACKUP_DIR / 'database')]
        if not backup_dirs:
            raise Exception('No backup directories found')
        date = max(backup_dirs)
    backup_dir = BACKUP_DIR / 'database' / date
    if not os.path.exists(backup_dir):
        raise Exception(f'There is no backup on {date}')
    for model in models:
        filename = f'{model.__name__.lower()}_backup.pkl'
        backup_path = backup_dir / filename
        with open(backup_path, 'rb') as file:
            backup = pickle.load(file)

        model.objects.bulk_create([obj for obj in backup])


def create_csv_backup():
    try:
        today = date.today()
        backup_dir = BACKUP_DIR / 'csv' / str(today)
        os.makedirs(backup_dir, exist_ok=True)

        backup_file = f'backup_{settings.EXCEL_FILE_NAME}'
        backup_path = backup_dir / backup_file

        shutil.copy2(settings.EXCEL_FILE_NAME, backup_path)

    except FileNotFoundError:
        raise FileNotFoundError(f"Original file not found: {settings.EXCEL_FILE_NAME}")

    except Exception as e:
        raise Exception(f"Error creating backup: {e}")


def create_csv_data():
    wb = Workbook()
    ws = wb.active
    ws.title = "Stan Techniczny Samochodow"

    for col_num, column_title in enumerate(settings.CSV_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.alignment = Alignment(horizontal="center")

    cars = Samochod.objects.all()
    for row_num in range(2, Samochod.objects.count() + 2):
        car = cars[row_num - 2]
        ws.cell(row=row_num, column=1, value=car.numer_rejestracyjny)
        ws.cell(row=row_num, column=2, value=random.randint(1000, 50000))
        ws.cell(row=row_num, column=3, value=generate_rating())
        ws.cell(row=row_num, column=4, value=generate_rating())
        ws.cell(row=row_num, column=5, value=generate_rating())
        ws.cell(row=row_num, column=6, value=generate_rating())
        ws.cell(row=row_num, column=7, value=generate_rating())
        ws.cell(row=row_num, column=8, value=generate_rating())
        ws.cell(row=row_num, column=9, value=random.randint(10, 1000))
        date_format = '%Y-%m-%d'
        start_date_10 = faker.date_between(start_date='-30d', end_date='-14d').strftime(date_format)
        start_date_11 = faker.date_between(start_date='-15d', end_date='today').strftime(date_format)

        ws.cell(row=row_num, column=10, value=start_date_10)
        ws.cell(row=row_num, column=11, value=start_date_11)

    wb.save(settings.EXCEL_FILE_NAME)

    with open(settings.CSV_FILE_NAME, 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            csvwriter.writerow(row)


def load_csv_backup(date=None):
    if date is None:
        backup_dirs = [d for d in os.listdir(BACKUP_DIR / 'csv')]
        if not backup_dirs:
            raise Exception('No backup directories found')
        date = max(backup_dirs)

    backup_dir = BACKUP_DIR / 'csv' / date
    backup_file = f'backup_{settings.EXCEL_FILE_NAME}'
    backup_path = backup_dir / backup_file
    if not os.path.exists(backup_path):
        raise FileNotFoundError(f'Backup file not found: {backup_path}')
    try:
        shutil.copy2(backup_path, settings.EXCEL_FILE_NAME)
    except FileNotFoundError as e:
        raise FileNotFoundError(f"Error loading backup: {e}")

    except Exception as e:
        raise Exception(f"Error loading backup: {e}")
